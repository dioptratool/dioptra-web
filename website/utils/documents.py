from decimal import Decimal

from django.conf import settings
from django.db.models import F, OuterRef, Subquery, Sum, Value
from django.db.models.functions import Coalesce
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import WHITE
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet

from website.currency import currency_name, currency_symbol
from website.models import Analysis, AnalysisCostType, FieldLabelOverrides
from website.models import InterventionInstance
from website.models.cost_line_item import CostLineItemInterventionAllocation
from website.models.cost_type import ProgramCost
from website.models.output_metric import OutputMetric
from website.workflows import AnalysisWorkflow

_gray_fill = PatternFill(start_color="00DADADA", end_color="00DADADA", fill_type="solid")

_dark_red_fill = PatternFill(start_color="00530000", end_color="00530000", fill_type="solid")

_black_side = Side(style="thin", color="000000")
_black_border = Border(left=_black_side, top=_black_side, right=_black_side, bottom=_black_side)


def _write_header_cell(ws, row, col, value):
    ws.cell(row=row, column=col, value=value)
    ws.cell(row=row, column=col).fill = _dark_red_fill
    ws.cell(row=row, column=col).font = Font(color=WHITE)
    ws.cell(row=row, column=col).border = _black_border


def _write_header_row(ws, row, values):
    for i, v in enumerate(values, start=1):
        _write_header_cell(ws, row=row, col=i, value=v)


def _get_other_cost_line_items(an_analysis: Analysis, intervention_instance: InterventionInstance):
    return (
        an_analysis.cost_line_items.filter(
            config__analysis_cost_type__in=[
                AnalysisCostType.IN_KIND,
                AnalysisCostType.CLIENT_TIME,
            ],
            config__allocations__intervention_instance=intervention_instance,
        )
        .annotate(coalesced_allocation=Coalesce(F("config__allocations__allocation"), Decimal("100.00")))
        .annotate(allocated_cost=F("total_cost") * (F("coalesced_allocation") / Value(100)))
        .order_by("config__analysis_cost_type", "budget_line_description")
    )


def _get_cost_breakdown_categories_and_cost_types(an_analysis: Analysis):
    return (
        an_analysis.cost_line_items.values(
            "config__cost_type__name",
            "config__category__name",
            "config__analysis_cost_type",
            "config__cost_type__type",
        )
        .annotate(
            allocated_cost_sum=Sum(F("total_cost") * (F("config__allocations__allocation") / Value(100)))
        )
        .order_by("-allocated_cost_sum")
    )


def _write_metadata_table(
    ws: worksheet,
    an_analysis: Analysis,
    intervention_instance: InterventionInstance,
    parameter_metadata: dict,
    analysis_url: str,
    starting_row: int = 1,
) -> int:
    """
    Write the metadata table to the provided worksheet in the upper left corner.

    Mutate the incoming parameter_metadata dictionary to contain row information for each parameter to be used later
    in the spreadsheet workflow

    Returns the last row with data on it to position other things on the page.
    """
    row = starting_row

    metadata = [
        ("Analysis Title", an_analysis.title),
        ("Analysis Type", getattr(an_analysis.analysis_type, "title", "")),
        ("Analysis Description", an_analysis.description),
        ("Analysis Start Date", an_analysis.start_date),
        ("Analysis End Date", an_analysis.end_date),
        ("Country", an_analysis.country.name),
        (
            FieldLabelOverrides.label_for("ci_grant_code", "Grants"),
            ", ".join(an_analysis.grants_list()),
        ),
        ("Intervention Being Analyzed", intervention_instance.display_name()),
    ]

    parameter_row = len(metadata) + 1
    for parameter_key, value in intervention_instance.parameters.items():
        for each_output_metric in intervention_instance.intervention.output_metric_objects():
            if parameter_key not in each_output_metric.parameters:
                continue
            if parameter_key in parameter_metadata:
                continue

            label = str(each_output_metric.parameters[parameter_key].label)
            metadata.append((label, value))

            # Store each metric's row for later use building formulas
            parameter_metadata[parameter_key] = parameter_row
            parameter_row += 1

    metadata += [
        ("Output count data source", an_analysis.output_count_source),
    ]

    author = ""
    if an_analysis.owner:
        author = an_analysis.owner.get_full_name()

    metadata += [
        ("Currency", f"{currency_name(analysis=an_analysis)}"),
        ("Owner", author),
        ("Analysis URL", analysis_url),
    ]

    for key, val in metadata:
        ws[f"A{row}"] = key
        ws[f"A{row}"].fill = _gray_fill
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = _black_border

        ws[f"B{row}"].fill = _gray_fill
        ws[f"B{row}"].border = _black_border
        ws[f"B{row}"].alignment = Alignment(horizontal="left")
        if key in [
            "Value of Cash Distributed",
            "Value of Business Grant Amount",
        ]:
            # For these values we format things as currency
            ws[f"B{row}"] = f"{currency_symbol(an_analysis)}{val:,.2f}"
        elif key in [
            "Analysis Start Date",
            "Analysis End Date",
        ]:
            # For these values we format things as a date
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = f"yyyy-MM-dd"
        else:
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = "#,##0.00"

        row += 1

    return row


def _write_cost_efficiency_table(
    ws: worksheet,
    an_analysis: Analysis,
    intervention_instance: InterventionInstance,
    starting_row: int = 1,
    metrics_all_costs_metadata: dict | None = None,
) -> int:
    """
    Write the cost efficiency table to the provided worksheet

    metrics_all_costs_metadata: This dict tracks the address for
      values that are referenced by excel functions in other parts of this
      sheet.

    Returns the last row with data on it to position other things on the page.

    """
    row = starting_row

    ws[f"A{row}"] = "Cost Efficiency"
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    intervention_metrics = intervention_instance.intervention.output_metric_objects()
    output_costs = an_analysis.output_costs[str(intervention_instance.id)]
    each_metric: OutputMetric
    for idx, each_metric in enumerate(intervention_metrics):
        if each_metric.id not in output_costs:
            continue
        ws[f"A{row}"] = f"{each_metric.metric_name} Program Costs only"
        ws[f"A{row}"].border = _black_border

        ws[f"B{row}"] = output_costs[each_metric.id]["direct_only"]
        ws[f"B{row}"].border = _black_border

        row += 1

        ws[f"A{row}"] = f"{each_metric.metric_name} including Program Costs, Support Costs, Indirect Costs"
        ws[f"A{row}"].border = _black_border

        metrics_all_costs_metadata[each_metric.metric_name] = f"B{row}"

        ws[f"B{row}"] = output_costs[each_metric.id]["all"]
        ws[f"B{row}"].border = _black_border

        row += 1

        if an_analysis.in_kind_contributions:
            ws[f"A{row}"] = (
                f"{each_metric.metric_name} including Program Costs, Support Costs, Indirect"
                f" Costs, In-Kind Contributions"
            )
            ws[f"A{row}"].border = _black_border

            ws[f"B{row}"] = output_costs[each_metric.id]["in_kind"]
            ws[f"B{row}"].border = _black_border

            row += 1

        if an_analysis.client_time and idx == (len(intervention_metrics) - 1):
            ws[f"A{row}"] = "Total Cost of Client Time"
            ws[f"A{row}"].border = _black_border

            ws[f"B{row}"] = output_costs[each_metric.id]["client"]
            ws[f"B{row}"].border = _black_border

            row += 1

    return row


def _write_cost_of_each_subcomponent_per_output_metric_table(
    ws: worksheet,
    an_analysis: Analysis,
    intervention_instance: InterventionInstance,
    starting_row: int,
) -> int:
    """
    Write the "Cost of Each Sub-component, per OUTPUT METRIC" table to the provided worksheet

    Returns the last row with data on it to position other things on the page.
    """

    row = starting_row

    for output_metric in intervention_instance.intervention.output_metric_objects():
        ws[f"A{row}"] = f"Cost of Each Sub-component, per {output_metric}"
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        percentages = an_analysis.subcomponent_cost_analysis.cost_line_item_average(
            exclude_support_costs=False
        )
        for idx, each_percentage in enumerate(percentages):
            ws[f"A{row}"] = f"{an_analysis.subcomponent_cost_analysis.subcomponent_labels[idx]}"
            ws[f"A{row}"].border = _black_border

            # This is a placeholder value that is used when building the
            # excel functions in `_fill_in_subcomponent_cost_efficiency_functions`
            ws[f"B{row}"] = str(output_metric.metric_name)
            ws[f"B{row}"].border = _black_border

            row += 1

    return row


def _write_cost_breakdown_table(
    ws: worksheet,
    an_analysis: Analysis,
    direct_cost_rows: list[int],
    starting_row: int,
) -> int:
    """
    Write the cost breakdown table to the provided worksheet.  The data in this table has
      functions that are dependent on other tables so this fills in blanks and then another
      function goes back through and fills them in.

    Mutate the incoming direct_cost_rows list to contain row information for each breakdown cost where
    Cost Type is a Program Cost

    Returns the last row with data on it to position other things on the page.
    """
    row = starting_row

    # Cost Breakdown Section
    ws[f"A{row}"] = "Cost Breakdown"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    _write_header_row(
        ws,
        row,
        [
            FieldLabelOverrides.label_for("ci_cost_type", "Cost Type"),
            "Category",
            "Amount",
            "% Of Total Amount",
        ],
    )

    row += 1

    first_data_row = row
    breakdown_data = _get_cost_breakdown_categories_and_cost_types(an_analysis)
    for each_breakdown in breakdown_data:
        cost_type_name = "Support"
        if each_breakdown["config__cost_type__name"]:
            cost_type_name = each_breakdown["config__cost_type__name"]

        category_name = "Other Supporting Costs"
        if each_breakdown["config__category__name"]:
            category_name = each_breakdown["config__category__name"]
        elif each_breakdown["config__analysis_cost_type"]:
            if each_breakdown["config__analysis_cost_type"] == AnalysisCostType.CLIENT_TIME:
                continue
            elif each_breakdown["config__analysis_cost_type"] == AnalysisCostType.IN_KIND:
                continue
            else:
                # This is the special case for Other HQ Costs
                category_name = AnalysisCostType.get_pretty_analysis_cost_type(
                    each_breakdown["config__analysis_cost_type"]
                )

        ws[f"A{row}"] = cost_type_name
        ws[f"A{row}"].border = _black_border

        ws[f"B{row}"] = category_name
        ws[f"B{row}"].border = _black_border

        ws[f"C{row}"] = 0
        ws[f"C{row}"].number_format = "#,##0.00"
        ws[f"C{row}"].border = _black_border

        ws[f"D{row}"] = 0
        ws[f"D{row}"].number_format = "#,##0.00"
        ws[f"D{row}"].border = _black_border

        # Store Program Cost rows necessary to input the proper formula
        if each_breakdown["config__cost_type__type"] == ProgramCost.id:
            direct_cost_rows.append(row)

        row += 1
    ws[f"C{row}"] = f"=SUM(C{first_data_row}:C{row - 1})"
    ws[f"C{row}"].number_format = "#,##0.00"
    ws[f"C{row}"].border = _black_border

    ws[f"D{row}"] = f"=SUM(D{first_data_row}:D{row - 1})"
    ws[f"D{row}"].number_format = "#,##0.00"
    ws[f"D{row}"].border = _black_border

    return row


def _fill_in_cost_breakdown_functions(
    ws: worksheet,
    first_data_row: int,
    last_data_row: int,
    full_cost_model_first_data_row: int,
    full_cost_model_last_data_row: int,
):
    row = first_data_row

    while row < last_data_row:
        ws[f"C{row}"] = (
            f"=SUMIFS("
            f"J{full_cost_model_first_data_row}:J{full_cost_model_last_data_row}, "
            f"A{full_cost_model_first_data_row}:A{full_cost_model_last_data_row}, "
            f"A{row}, "
            f"B{full_cost_model_first_data_row}:B{full_cost_model_last_data_row}, "
            f"B{row}"
            f")"
        )
        ws[f"D{row}"] = (
            f"=C{row}/SUM(J{full_cost_model_first_data_row}:J{full_cost_model_last_data_row}) * 100"
        )
        row += 1


def _fill_in_cost_efficiency_functions(
    ws: worksheet,
    an_analysis: Analysis,
    intervention_instance: InterventionInstance,
    parameter_metadata: dict,
    direct_cost_rows: list[int],
    other_cost_rows: dict[int, list[int]],
    full_cost_row: int,
    efficiency_row: int,
):
    """
    Retroactively update the Cost Efficiency rows with formula values
    """

    # Convert list of direct cost rows to an Excel-compatible string (all will be in column C)
    direct_cost_ids = ", ".join([f"C{row}" for row in direct_cost_rows])

    # Convert list of other cost rows to an Excel-compatible string (all will be in column C)
    in_kind_ids = ", ".join([f"E{row}" for row in other_cost_rows[int(AnalysisCostType.IN_KIND)]])
    client_time_ids = ", ".join([f"E{row}" for row in other_cost_rows[int(AnalysisCostType.CLIENT_TIME)]])

    # For each Metric, convert its calculate method to an Excel-compatible string
    an_analysis_metrics = intervention_instance.intervention.output_metric_objects()
    for idx, metric in enumerate(an_analysis_metrics):
        param_to_excel_map = {}
        for param_name in metric.parameters:
            meta_row = parameter_metadata.get(param_name)
            if not meta_row:
                continue
            param_to_excel_map[param_name] = f"B{meta_row}"

        # Direct Cost only
        param_to_excel_map.update({"cost_output_sum": f"SUM({direct_cost_ids})"})
        eq_1 = metric.convert_calculate_to_excel_formula(param_to_excel_map)
        if eq_1:
            ws[f"B{efficiency_row}"] = f"=FIXED({eq_1}, 2)"
            efficiency_row += 1

        # Full Cost
        param_to_excel_map.update({"cost_output_sum": f"C{full_cost_row}"})
        eq_2 = metric.convert_calculate_to_excel_formula(param_to_excel_map)
        if eq_2:
            ws[f"B{efficiency_row}"] = f"=FIXED({eq_2}, 2)"
            efficiency_row += 1

        # In-Kind Cost
        # This value should be added on top of the above full_cost_row
        if an_analysis.in_kind_contributions:
            param_to_excel_map.update({"cost_output_sum": f"SUM({in_kind_ids})"})
            eq_3 = metric.convert_calculate_to_excel_formula(param_to_excel_map)
            if eq_3:
                ws[f"B{efficiency_row}"] = f"=FIXED(({eq_2}) + ({eq_3}), 2)"
                efficiency_row += 1

        # Client Time Cost
        # This value is a standard sum and not calculated on a "per unit" basis
        if an_analysis.client_time and idx == (len(an_analysis_metrics) - 1):
            ws[f"B{efficiency_row}"] = f"=FIXED(SUM({client_time_ids}), 2)"
            efficiency_row += 1


def _fill_in_subcomponent_cost_efficiency_functions(
    ws: worksheet,
    metrics_all_costs_metadata: dict,
    first_data_row: int,
    last_data_row: int,
    full_cost_model_first_data_row: int,
    full_cost_model_last_data_row: int,
):
    """
    The sum of the subcomponent costs / the Item Total

    This is a complex function needing to know the location of a number of dynamic values.
    """

    row = first_data_row

    # Get a list of strings that are the header values.   This is used to lookup a value used in the excel function
    full_cost_model_headers = [c.value for c in ws[full_cost_model_first_data_row - 1]]

    while row < last_data_row:
        # Find which column has the Subcomponent Total for the Subcomponent in this Row
        try:
            subcomponent_column = get_column_letter(
                full_cost_model_headers.index(ws[f"A{row}"].value + " Total") + 1
            )
        except ValueError:
            # We can assume when this happens it has hit a header for one of the Subcomponent Sub Sections.
            # Something that looks like the following.  Where "Number of People" and "Number of Days of Training" are
            # the two Output Metrics for an intervention.  The ValueError is raised when it hits A27 which
            # doesn't have a value to lookup in the subcomponent label list.  We can just skip it
            # since the value is blank on that row.   We are filling in column B for this sub section.
            #
            # Example layout:
            #
            #   Row#    A
            #   24      Cost of Each Sub-component, per Number of People
            #   25      subcomponentlabel1
            #   26      subcomponentlabel2
            #   27      Cost of Each Sub-component, per Number of Days of Training
            #   28      subcomponentlabel1
            #   29      subcomponentlabel2
            #
            row += 1
            continue
        # Create a function that takes the percentage of the Subcomponent to the relevant Cost Total
        #   and applies that percentage to relevant metric's Efficiency Cost
        if ws[f"B{row}"].value in metrics_all_costs_metadata:
            v = (
                "="
                + metrics_all_costs_metadata[ws[f"B{row}"].value]
                + " * "
                + f"SUM({subcomponent_column}{full_cost_model_first_data_row}:{subcomponent_column}{full_cost_model_last_data_row})"
                + " / "
                + f"SUMIFS("
                f"J{full_cost_model_first_data_row}:J{full_cost_model_last_data_row},"
                f"{subcomponent_column}{full_cost_model_first_data_row}:{subcomponent_column}{full_cost_model_last_data_row}, "
                '"<>"'
                f")"
            )
        else:
            v = ""
        c = ws.cell(
            row=row,
            column=2,
            value=v,
        )

        c.number_format = "$#,##0.00"

        row += 1


def _write_full_cost_model_table(
    ws: worksheet,
    an_analysis: Analysis,
    starting_row: int,
    intervention_instance: InterventionInstance,
) -> int:
    """
    Write the full cost model table to the provided worksheet

    Returns the last row with data on it to position other things on the page.
    """
    row = starting_row

    ws[f"A{row}"] = "Cost Model"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    header_row = [
        "Cost Type",
        "Category",
        "Cost Item",
        FieldLabelOverrides.label_for("ci_grant_code", "Grant"),
        "Sector Code",
        FieldLabelOverrides.label_for("tr_site_code", "Site"),
        FieldLabelOverrides.label_for("ci_total_cost", "Total Cost"),
        "% to Intervention",
        "Notes",
        "Item Total",
    ]

    if (
        hasattr(an_analysis, "subcomponent_cost_analysis")
        and an_analysis.subcomponent_cost_analysis.subcomponent_labels_confirmed
    ):
        for each_label in an_analysis.subcomponent_cost_analysis.subcomponent_labels:
            header_row.append(each_label)
            header_row.append(each_label + " Total")

    _write_header_row(ws, row, header_row)

    row += 1

    cost_line_items = (
        an_analysis.cost_line_items.all()
        .exclude(config__analysis_cost_type__in=[AnalysisCostType.IN_KIND, AnalysisCostType.CLIENT_TIME])
        .prefetch_related(
            "config",
            "config__cost_type",
            "config__category",
            "config__allocations",
        )
        .annotate(
            allocated_cost=F("total_cost")
            * (
                Subquery(
                    CostLineItemInterventionAllocation.objects.filter(
                        cli_config=OuterRef("config"),
                        intervention_instance=intervention_instance,
                    ).values("allocation")[:1]
                )
                / Value(100)
            )
        )
        .order_by("-allocated_cost")
    )

    for each_cost_line_item in cost_line_items:
        cost_type_name = "Support"
        if each_cost_line_item.config.cost_type:
            cost_type_name = each_cost_line_item.config.cost_type.name
        elif each_cost_line_item.config.analysis_cost_type:
            cost_type_name = each_cost_line_item.config.get_pretty_analysis_cost_type

        category_name = "Other Supporting Costs"
        if each_cost_line_item.config.category:
            category_name = each_cost_line_item.config.category.name
        elif each_cost_line_item.config.analysis_cost_type:
            category_name = each_cost_line_item.config.get_pretty_analysis_cost_type

        # Use prefetch cache by iterating instead of .get() which bypasses cache
        cli_allocation = 0
        for allocation in each_cost_line_item.config.allocations.all():
            if allocation.intervention_instance_id == intervention_instance.id:
                cli_allocation = allocation.allocation
                break

        row_data = [
            cost_type_name,
            category_name,
            each_cost_line_item.budget_line_description,
            each_cost_line_item.grant_code,
            each_cost_line_item.sector_code,
            each_cost_line_item.site_code,
            each_cost_line_item.total_cost,
            cli_allocation,
            each_cost_line_item.note,
            f"=(G{row} * H{row})/100",
        ]
        for i, each in enumerate(row_data, start=1):
            ws.cell(row=row, column=i, value=each)

        ws[f"G{row}"].number_format = "#,##0.00"
        ws[f"H{row}"].number_format = "#,##0.00"
        ws[f"J{row}"].number_format = "#,##0.00"

        if (
            hasattr(an_analysis, "subcomponent_cost_analysis")
            and an_analysis.subcomponent_cost_analysis.subcomponent_labels_confirmed
            and getattr(each_cost_line_item.config, "subcomponent_analysis_allocations", None)
        ):
            subcomponent_analysis_start_column = len(row_data) + 1
            for (
                idx,
                each_subcomponent_allocation,
            ) in each_cost_line_item.config.subcomponent_analysis_allocations.items():
                percentage = ws.cell(
                    row=row,
                    column=subcomponent_analysis_start_column + (int(idx) * 2),
                    value=Decimal(each_subcomponent_allocation) / 100,
                )
                total = ws.cell(
                    row=row,
                    column=subcomponent_analysis_start_column + ((int(idx) * 2) + 1),
                    value=f"={get_column_letter(len(row_data))}{row} * {get_column_letter(subcomponent_analysis_start_column + (int(idx) * 2))}{row}",
                )
                percentage.number_format = "0.00%"
                total.number_format = "#,##0.00"

        row += 1

    return row


def _write_other_cost_model_table(
    ws: worksheet,
    an_analysis: Analysis,
    intervention_instance: InterventionInstance,
    other_cost_rows: dict[int, list[int]],
    starting_row: int,
) -> int:
    """
    Write the other cost model table to the provided worksheet

    Returns the last row with data on it to position other things on the page.
    """
    row = starting_row

    # Early exit if there are no other costs for In-Kind Contributions or Client Time
    # Other HQ Costs are included in the standard cost model table
    if not (an_analysis.client_time or an_analysis.in_kind_contributions):
        return row

    ws[f"A{row}"] = "Other HQ Costs"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    _write_header_row(
        ws,
        row,
        [
            "Category",
            "Cost Item",
            "Total Cost",
            "% of Intervention",
            "Item Total",
            "Notes",
        ],
    )

    row += 1

    for each_cost_line_item in _get_other_cost_line_items(an_analysis, intervention_instance):
        ws[f"A{row}"] = each_cost_line_item.config.get_pretty_analysis_cost_type
        ws[f"B{row}"] = each_cost_line_item.budget_line_description
        ws[f"C{row}"] = each_cost_line_item.total_cost
        ws[f"C{row}"].number_format = "#,##0.00"
        ws[f"D{row}"] = each_cost_line_item.coalesced_allocation
        ws[f"D{row}"].number_format = "#,##0.00"
        ws[f"E{row}"] = f"=(C{row} * D{row})/100"
        ws[f"E{row}"].number_format = "#,##0.00"
        ws[f"F{row}"] = each_cost_line_item.note

        cost_type = int(each_cost_line_item.config.analysis_cost_type)
        if cost_type in other_cost_rows:
            other_cost_rows[cost_type].append(row)

        row += 1

    return row


def _formatting_pass(ws: worksheet):
    for i, each_col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 26
