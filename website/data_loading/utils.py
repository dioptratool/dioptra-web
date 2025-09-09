import contextlib
import csv
import io
import logging
from collections import Counter
from decimal import Decimal
from typing import AnyStr, IO

import xlrd
from openpyxl import load_workbook

from website.models.cost_line_item import CostLineItem

logger = logging.getLogger(__name__)


################################################################################
# Cost line items from upload.
################################################################################


def excel_file_to_array(f: IO[AnyStr]) -> list[list[str]]:
    """
    Read either .xls or .xlsx from any file-like, return a list of non-empty
    rows with every cell coerced to str.
    """
    # read whole file into memory
    data = f.read()
    buf = io.BytesIO(data)

    if is_xlsx(buf):
        wb = load_workbook(buf, data_only=True)
        raw_rows = [[cell.value for cell in row] for row in wb.active.iter_rows()]

    elif is_xls(buf):
        wb = xlrd.open_workbook(file_contents=data)
        sheet = wb.sheet_by_index(0)
        raw_rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    else:
        # decode bytes → text, trying UTF-8 then CP1252
        if isinstance(data, (bytes, bytearray)):
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("cp1252")
        else:
            text = data  # already str

        # split into lines and parse
        reader = csv.reader(text.splitlines())
        raw_rows = [row for row in reader]

    # drop totally empty rows, then stringify every cell
    cleaned = []
    for row in raw_rows:
        if not any(cell not in (None, "") for cell in row):
            continue
        cleaned.append([str(cell) if cell is not None else "" for cell in row])

    return cleaned


def excel_file_to_dict(f: IO[AnyStr], normalize_headers=True) -> list[dict]:
    """
    This is for parsing files that we expect to have a header row.
    """
    dict_list = []
    data = excel_file_to_array(f)
    # read first row for keys
    keys = data[0]

    if normalize_headers:
        keys = [k.lower().replace(" ", "_").replace(",", "") for k in keys]
    values = [data[i] for i in range(1, len(data))]

    for value in values:
        dict_list.append(dict(zip(keys, value)))
    return dict_list


def cast_to_decimal_four_decimal_places(value):
    return round(Decimal(value), 4)


def cast_dollar_to_decimal_four_decimal_places(value):
    return cast_to_decimal_four_decimal_places(value.replace("$", "").replace(",", ""))


def cast_and_handle_numeric_strings(value):
    """
    Some values like Grant Code can be 100 but excel treats this
    like a number, and it comes in as a float (i.e., 100.00).

    Here we try catch those occurrences
    """
    if value is None:
        return value
    try:
        value = str(int(float(value)))
    except ValueError:
        pass

    return value


def cast_boolean_type(value) -> bool:
    if not value or value.lower() in ["false", "no", "", "0"]:
        return False
    elif value.lower() in ["true", "yes", "1"]:
        return True
    else:
        raise ValueError("Invalid boolean value")


def _human_field_name(field_name):
    if not hasattr(_human_field_name, "_name_map"):
        _human_field_name._name_map = {field.name: field.verbose_name for field in CostLineItem._meta.fields}
    return _human_field_name._name_map.get(field_name, field_name)


################################################################################
# Transactions from data store.
################################################################################


class BulkInserter(contextlib.AbstractContextManager):
    """
    The character used to represent a null value. If we use the default,
    we get empty strings converted to NULLs, which we don't want.
    But in some cases, we do want NULLs- we can replace their None value in a CSV or whatever string
    with this control character.
    """

    null = "\u0001"

    def __init__(self, conn, table, debug=False):
        self.debug = debug
        self.rows = 0
        self.conn = conn
        self.table = table
        self.stream = io.StringIO()
        self.writer = None

    def add_row(self, row: dict):
        self.rows += 1
        if self.writer is None:
            self.writer = csv.DictWriter(self.stream, delimiter="\t", fieldnames=row.keys())
            self.writer.writeheader()
        self.writer.writerow(row)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        if self.writer is not None:
            self.stream.flush()
            self.stream.seek(0)
            if self.debug:
                print(f"Inserting {self.rows} rows")
                print(self.stream.getvalue())
                self.stream.seek(0)
            with self.conn.cursor() as cursor:
                fieldnames = [f'"{f}"' for f in self.writer.fieldnames]
                fieldnames = ", ".join(fieldnames)
                sql = (
                    f"COPY {self.table} ({fieldnames}) "
                    # We do not want nulls, so instead of an empty string being turned to NULL,
                    # use something that won't match anything.
                    f"FROM STDIN WITH HEADER DELIMITER '\t' NULL '{self.null}' CSV"
                )

                with cursor.copy(sql) as copy:
                    chunk = self.stream.read(8192)
                    # continue until EOF (read() returns b'' or '')
                    while chunk:
                        copy.write(chunk)
                        chunk = self.stream.read(8192)
        return super().__exit__(exc_type, exc_value, traceback)


########
# Misc
########


def get_duplicates(lst):
    return [item for item, count in Counter(lst).items() if count > 1]


def is_xls(uploaded_file: IO[AnyStr]) -> bool:
    OLE_HEADER = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    # read just enough to check the header

    header = uploaded_file.read(len(OLE_HEADER))
    uploaded_file.seek(0)
    return header == OLE_HEADER


def is_xlsx(uploaded_file: IO[AnyStr]) -> bool:
    # read just enough to check the header
    start = uploaded_file.read(4)
    uploaded_file.seek(0)
    return start[:2] == b"PK"  # ZIP files start with “PK”
